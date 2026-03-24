import openpyxl

try:
    wb = openpyxl.load_workbook('FIVE.xlsx', data_only=True)
    if 'matchs_joueurs' not in wb.sheetnames:
        print("Sheet 'matchs_joueurs' not found. Available sheets:", wb.sheetnames)
        exit()

    sheet = wb['matchs_joueurs']
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    print("Headers:", headers)

    # Get first 5 rows of data
    print("First 5 rows:")
    for i in range(2, 7):
        print(f"Row {i}:", [cell.value for cell in sheet[i]])
    
except Exception as e:
    print(f"Error: {e}")
