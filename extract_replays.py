import openpyxl
import json
import datetime

def extract_replays():
    try:
        wb = openpyxl.load_workbook('FIVE.xlsx', data_only=True)
        if 'videos' not in wb.sheetnames:
            print("Sheet 'videos' not found.")
            return

        sheet = wb['videos']
        replays = []
        
        # Headers: date (A), Equipe A (B), Score A (C), Equipe B (D), Score B (E), Lien (F), timeline (G)
        # 0-indexed: 0, 1, 2, 3, 4, 5, 6
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_val = row[0]
            if not date_val: continue
            
            # Format Date
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%d/%m/%Y')
            else:
                date_str = str(date_val)
                
            team_a_str = row[1]
            score_a = row[2]
            team_b_str = row[3]
            score_b = row[4]
            link = row[5]
            timeline_raw = row[6]
            
            # Process Teams
            team_a = [t.strip() for t in str(team_a_str).split(';')] if team_a_str else []
            team_b = [t.strip() for t in str(team_b_str).split(';')] if team_b_str else []
            
            # Process Timeline
            timeline = []
            if timeline_raw:
                try:
                    # Format: "4,Jesus,0-1;6,Enos,1-1;..."
                    events = str(timeline_raw).split(';')
                    for event in events:
                        parts = event.split(',')
                        if len(parts) >= 3:
                            timeline.append({
                                "minute": parts[0].strip(),
                                "scorer": parts[1].strip(),
                                "score": parts[2].strip()
                            })
                except Exception as e:
                    print(f"Error parsing timeline for {date_str}: {e}")

            match_data = {
                "date": date_str,
                "teamA": team_a,
                "scoreA": score_a,
                "teamB": team_b,
                "scoreB": score_b,
                "link": link,
                "timeline": timeline
            }
            replays.append(match_data)
            
        # Serialize to JS
        js_content = f"window.replays = {json.dumps(replays, indent=4)};"
        
        with open('js/replays_data.js', 'w') as f:
            f.write(js_content)
            
        print(f"Successfully extracted {len(replays)} replays to js/replays_data.js")
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_replays()
