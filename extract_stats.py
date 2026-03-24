import openpyxl
import json
import os

def extract_from_excel():
    try:
        wb = openpyxl.load_workbook('FIVE.xlsx', data_only=True)
        if 'statistiques' not in wb.sheetnames:
            print("Sheet 'statistiques' not found.")
            return

        sheet = wb['statistiques']
        
        # Headers are in row 1, data starts row 2
        # Columns based on inspection:
        # A: player
        # B: participations
        # C: wins
        # D: nuls
        # E: goals_total
        # F: assists_total
        # G: avg_time_to_goal
        # H: meilleur buteur
        # I: meilleur passeur
        
        html_output = ""
        
        
        def safe_float(val):
            try:
                if val is None: return 0.0
                if isinstance(val, str):
                   val = val.replace(',', '.') # Handle comma decimals if any
                   # Handle "-" or empty
                   if val.strip() in ['-', '', ' ']: return 0.0
                return float(val)
            except:
                return 0.0

        # Use a dict to store history per player: {player_name: [{date: "...", goals: N, assists: N}, ...]}
        history_data = {}
        
        if 'matchs_joueurs' in wb.sheetnames:
            history_sheet = wb['matchs_joueurs']
            # Headers: date (A), player (B), goals (C), assists (D) -> indices 0, 1, 2, 3
            
            for row in history_sheet.iter_rows(min_row=2, values_only=True):
                # Check row validity
                if not row[0] or not row[1]: continue
                
                date_val = row[0]
                player_name = row[1]
                goals_val = safe_float(row[2])
                assists_val = safe_float(row[3])
                
                # Format date to DD/MM
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%d/%m')
                else:
                    date_str = str(date_val)
                    
                entry = {
                    "date": date_str,
                    "goals": int(goals_val),
                    "assists": int(assists_val)
                }
                
                if player_name not in history_data:
                    history_data[player_name] = []
                history_data[player_name].append(entry)

        # Iterate rows starting from 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name: continue
            
            matchs = safe_float(row[1])
            wins = safe_float(row[2])
            nuls = safe_float(row[3])
            goals = safe_float(row[4])
            assists = safe_float(row[5])
            tmb = safe_float(row[6])
            best_scorer = safe_float(row[7])
            best_passer = safe_float(row[8])
            
            # Format filename
            fname = name.lower().replace(" ", "_") + ".png"
            
            # Recalculate TMB correctly as "Minutes per Goal Involvement" (Integer)
            # User formula inverted for time: (Matches * 60) / (Total Goal Involvements)
            total_inv = goals + assists
            if total_inv > 0 and matchs > 0:
                tmb_val = (matchs * 60) / total_inv
                tmb = f"{int(round(tmb_val))}" # Integer string
            else:
                tmb = "0"
            
            # Generate History JSON string safely
            # Use data attribute compliant JSON (replace " with &quot;)
            if name in history_data:
                # Sort by date ? Date is string DD/MM, might be tricky if years differ, but usually OK for single season. 
                # Assuming Excel is chronologically ordered is safer.
                phistory = history_data[name]
                history_json = json.dumps(phistory).replace('"', '&quot;')
            else:
                history_json = "[]"
            
            # Generate HTML item
            html_count = len(history_data.get(name, []))
            
            # Legacy logic (simplified for extraction)
            losses = matchs - wins - nuls
            if losses < 0: losses = 0
            
            # Convert to int explicitly for display requirements
            item_html = f'''
                                <div class="player-list-item" 
                                     data-name="{name}" 
                                     data-fname="{fname}"
                                     data-matchs="{int(matchs)}" 
                                     data-wins="{int(wins)}" 
                                     data-draws="{int(nuls)}" 
                                     data-losses="{int(losses)}" 
                                     data-goals="{int(goals)}" 
                                     data-assists="{int(assists)}" 
                                     data-tmb="{tmb}"
                                     data-best-scorer="{int(best_scorer)}"
                                     data-best-passer="{int(best_passer)}">
                                     <div class="avatar-sm" style="background-image: url('joueurs/{fname}'); background-size: cover;"></div>
                                     <div class="player-list-info"><span class="name">{name}</span></div>
                                </div>'''
            html_output += item_html

        print(html_output)
        
        # Also print the GLOBAL JSON
        print("\n\n--- GLOBAL HISTORY JSON ---")
        print(f"const playerHistory = {json.dumps(history_data, indent=4)};")
        
        with open('extracted_list.html', 'w') as f:
            f.write(html_output)
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_from_excel()
