
import openpyxl

def inspect_videos():
    try:
        wb = openpyxl.load_workbook('FIVE.xlsx', data_only=True)
        if 'videos' not in wb.sheetnames:
            print("Sheet 'videos' not found.")
            return

        sheet = wb['videos']
        
        print("Headers (Row 1):")
        headers = [cell.value for cell in sheet[1]]
        print(headers)
        
        print("\nFirst 3 rows of data:")
        for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
            print(row)
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_videos()
